import openpyxl
from openpyxl.styles import PatternFill

def calcular_espaciado_columnas(ws):
    """
    Calcula el espaciado entre columnas de una hoja de cálculo.

    Args:
        ws (Worksheet): La hoja de cálculo sobre la cual se realizará el cálculo.

    Returns:
        dict: Un diccionario con el espaciado entre columnas, donde las claves son las letras de las columnas
        y los valores son los anchos respectivos.
    """
    # Obtener el número de columnas con datos
    num_columnas = ws.max_column

    # Calcular el ancho para cada columna
    ancho_base = 20
    espaciado_entre_columnas = {}
    for col in range(1, num_columnas + 1):
        letra_columna = openpyxl.utils.get_column_letter(col)
        espaciado_entre_columnas[letra_columna] = ancho_base

    return espaciado_entre_columnas
def aplicar_estilo_excel(nombre_archivo):
    """
    Aplica un estilo específico a un archivo de Excel, incluyendo espaciado entre columnas, color al encabezado,
    y fijación del encabezado.

    Args:
        nombre_archivo (str): La ruta del archivo de Excel al cual se le aplicará el estilo.
    """
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(nombre_archivo)

    # Iterar sobre todas las hojas del libro de trabajo
    for ws in wb:
        # Calcular el espaciado entre columnas
        espaciado_entre_columnas = calcular_espaciado_columnas(ws)

        # Aplicar el espaciado entre columnas
        for col, width in espaciado_entre_columnas.items():
            ws.column_dimensions[col].width = width

        # Dar color al encabezado
        encabezado_fill = PatternFill(start_color="FFC305", end_color="FFC305", fill_type="solid")
        for cell in ws[1]:
            cell.fill = encabezado_fill

        # Fijar el encabezado
        ws.freeze_panes = 'A2'

        ws.auto_filter.ref = ws.dimensions

    # Guardar los cambios en el mismo archivo Excel
    wb.save(nombre_archivo)

    print(f"Se ha aplicado el estilo al archivo {nombre_archivo}")
